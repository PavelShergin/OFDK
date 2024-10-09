import openpyxl
from django.contrib import admin
from .models import Order, OrderItem
from django.http import HttpResponse
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side


class OrderItemInline(admin.TabularInline):
    model = OrderItem
    raw_id_fields = ['product']


class OrderAdmin(admin.ModelAdmin):
    list_display = ['id', 'first_name', 'last_name', 'address', 'phone_number', 'paid', 'created', 'updated']
    list_filter = ['paid', 'created', 'updated']
    inlines = [OrderItemInline]
    exclude = ['basket_history']

    def generate_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.merge_cells('A1:E1')
        ws[
            'A1'] = f'{queryset.first().first_name} {queryset.first().last_name}, {queryset.first().address}, {queryset.first().phone_number}'

        currency_symbol = '₽'

        for order in queryset:
            ws.append([order.created.replace(tzinfo=None), '', '', '', ''])

            ws.append(['Наименование', 'Цена', 'Кол-во', 'Ед. изм.', 'Стоимость'])

            total_cost = 0
            for item in order.items.all():
                item_cost = str(item.get_cost()) + currency_symbol
                total_cost += item.get_cost()
                parent_category_product = f'{item.product.category.parent}/{item.product.name}'
                ws.append(
                    [parent_category_product, f'{item.price} {currency_symbol}', item.quantity,
                     item.product.unit_of_measurement, item_cost])

            ws.append(['', '', '', 'Итого:', f'{total_cost} {currency_symbol}'])

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(
                    horizontal='right' if cell.column > 1 else 'left')
                cell.border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))

        for column in ws.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            ws.column_dimensions[column_letter].width = adjusted_width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="order_items.xlsx"'
        response.write(buffer.getvalue())

        return response

    generate_excel.short_description = "Сгенерировать Excel"
    actions = ['generate_excel']


admin.site.register(Order, OrderAdmin)
